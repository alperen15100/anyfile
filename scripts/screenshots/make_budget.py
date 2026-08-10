#!/usr/bin/env python3
"""
Sample multi-sheet workbook for the Gander screenshots. Entirely fictional.

Values are written as literals rather than formulas on purpose: SheetJS reads
the cached result stored next to a formula, and openpyxl does not write one, so
every computed cell would come out blank in the app.
"""
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "out"
OUT.mkdir(exist_ok=True)

rnd = random.Random(20260808)

HDR_FILL = PatternFill("solid", fgColor="1F3A5F")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="B7C3D2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "#,##0"


def header_row(ws, row, ncols, height=30):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left",
                                   vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = height


SECTIONS = [
    ("STAFF COSTS", [
        ("4100 · Salaries and wages", 184200),
        ("4110 · Employer national insurance", 25420),
        ("4115 · Employer pension", 11060),
        ("4120 · Overtime and standby", 14800),
        ("4130 · Agency cover", 42800),
        ("4140 · Apprenticeship levy", 920),
    ]),
    ("PROPERTY AND DEPOTS", [
        ("4300 · Depot rent", 21400),
        ("4305 · Business rates", 6050),
        ("4310 · Electricity", 3980),
        ("4315 · Gas", 1140),
        ("4320 · Water and wastewater", 860),
        ("4325 · Waste and recycling", 2210),
        ("4330 · Cleaning and grounds", 1780),
        ("4335 · Building maintenance", 4620),
    ]),
    ("FLEET", [
        ("4400 · Vehicle lease", 19400),
        ("4405 · Fuel", 14500),
        ("4410 · Servicing and repair", 9800),
        ("4415 · Tyres", 2140),
        ("4420 · Motor insurance", 5300),
        ("4425 · Telematics subscription", 1180),
    ]),
    ("OPERATIONS", [
        ("4500 · Plant and equipment hire", 18600),
        ("4505 · Small tools", 3240),
        ("4510 · Consumables", 7180),
        ("4515 · PPE and workwear", 4070),
        ("4520 · Laboratory analysis", 6350),
        ("4525 · Waste disposal", 5940),
        ("4530 · Subcontract survey", 12800),
    ]),
    ("TRAVEL AND TRAINING", [
        ("4600 · Travel and mileage", 5900),
        ("4610 · Accommodation and subsistence", 2500),
        ("4620 · Training courses", 4800),
        ("4630 · Certification and renewals", 2800),
    ]),
    ("TECHNOLOGY", [
        ("4700 · Software and licences", 14750),
        ("4710 · IT hardware", 5200),
        ("4720 · Mobile and connectivity", 2380),
    ]),
    ("OVERHEADS", [
        ("4800 · Marketing", 4300),
        ("4810 · Tender and bid costs", 8000),
        ("4900 · Liability insurance", 16400),
        ("4910 · Professional fees", 9100),
        ("4920 · Bank and card charges", 1450),
        ("4930 · Audit and accountancy", 3600),
        ("4990 · Contingency", 15000),
    ]),
]


def summary_sheet(wb):
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Q3 FY26 Operating Budget — Consolidated"
    ws["A1"].font = Font(bold=True, size=14, color="1F3A5F")
    ws["A2"] = ("Harrowgate Field Services Ltd  ·  prepared 8 Aug 2026  ·  "
                "all figures GBP  ·  Sep is forecast")
    ws["A2"].font = Font(size=9, italic=True, color="666666")

    cols = ["Cost centre", "Jul budget", "Jul actual", "Aug budget", "Aug actual",
            "Sep budget", "Sep forecast", "Q3 budget", "Q3 forecast", "Variance",
            "Var %"]
    ws.append([])
    ws.append(cols)
    header_row(ws, 4, len(cols))

    r = 5
    grand = [0] * 8
    for title, items in SECTIONS:
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = Font(bold=True, size=10, color="1F3A5F")
        r += 1
        sect = [0] * 8
        for name, base in items:
            jb = base
            ja = int(base * rnd.uniform(0.90, 1.14))
            ab = base if rnd.random() < 0.75 else int(base * rnd.uniform(0.85, 1.2))
            aa = int(ab * rnd.uniform(0.88, 1.16))
            sb = base if rnd.random() < 0.7 else int(base * rnd.uniform(0.9, 1.15))
            sf = int(sb * rnd.uniform(0.93, 1.11))
            qb, qf = jb + ab + sb, ja + aa + sf
            var = qb - qf
            vals = [jb, ja, ab, aa, sb, sf, qb, qf]
            for i, v in enumerate(vals):
                sect[i] += v
            ws.cell(row=r, column=1, value=name)
            for i, v in enumerate(vals):
                ws.cell(row=r, column=2 + i, value=v)
            ws.cell(row=r, column=10, value=var)
            ws.cell(row=r, column=11, value=(var / qb) if qb else 0)
            r += 1

        ws.cell(row=r, column=1, value=f"   {title.title()} subtotal").font = Font(bold=True)
        for i, v in enumerate(sect):
            ws.cell(row=r, column=2 + i, value=v).font = Font(bold=True)
        sv = sect[6] - sect[7]
        ws.cell(row=r, column=10, value=sv).font = Font(bold=True)
        ws.cell(row=r, column=11, value=sv / sect[6]).font = Font(bold=True)
        for i in range(8):
            grand[i] += sect[i]
        r += 1

    ws.cell(row=r, column=1, value="TOTAL OPERATING COST").font = Font(bold=True, size=11)
    for i, v in enumerate(grand):
        ws.cell(row=r, column=2 + i, value=v).font = Font(bold=True, size=11)
    gv = grand[6] - grand[7]
    ws.cell(row=r, column=10, value=gv).font = Font(bold=True, size=11)
    ws.cell(row=r, column=11, value=gv / grand[6]).font = Font(bold=True, size=11)

    for rr in range(5, r + 1):
        for c in range(1, 12):
            cell = ws.cell(row=rr, column=c)
            cell.border = BORDER
            if 2 <= c <= 10:
                cell.number_format = MONEY
            if c == 11:
                cell.number_format = "0.0%"

    ws.column_dimensions["A"].width = 34
    for c in range(2, 12):
        ws.column_dimensions[get_column_letter(c)].width = 12.5
    ws.freeze_panes = "B5"
    return ws


def depots_sheet(wb):
    ws = wb.create_sheet("Depots")
    ws["A1"] = "Cost by depot — Q3 FY26"
    ws["A1"].font = Font(bold=True, size=13, color="1F3A5F")
    hdr = ["Depot", "Region", "Headcount", "Jul", "Aug", "Sep (f)", "Q3 total",
           "Cost per head", "vs Q2", "Status"]
    ws.append([])
    ws.append(hdr)
    header_row(ws, 3, len(hdr))
    depots = [
        ("Willowmere Yard", "North", 34, 96420, 99180, 94300, 0.031, "On plan"),
        ("Kestrel Bay Depot", "Coastal", 28, 81340, 88705, 83900, 0.094, "Watch"),
        ("Thornhill Row", "North", 19, 54860, 52190, 55400, -0.012, "On plan"),
        ("Marrow Lane Works", "Central", 41, 118900, 121460, 116750, 0.058, "Watch"),
        ("Fenwick Sidings", "Central", 12, 33710, 31085, 34200, -0.024, "On plan"),
        ("Alderbrook Store", "South", 9, 21480, 22940, 21900, 0.017, "On plan"),
        ("Greywater Pumping", "Coastal", 15, 44290, 47630, 45100, 0.122, "Over"),
        ("Saltmarsh Access", "Coastal", 7, 18960, 17240, 19400, -0.036, "On plan"),
        ("Hollowfield Compound", "South", 23, 66180, 64920, 67300, 0.008, "On plan"),
        ("Brackenridge Sub-depot", "North", 11, 29740, 33810, 30200, 0.071, "Watch"),
        ("Coldharbour Annexe", "Central", 6, 15220, 14680, 15600, -0.019, "On plan"),
        ("Pennyfields Store", "South", 8, 19870, 21050, 20100, 0.042, "On plan"),
    ]
    r = 4
    tot = [0, 0, 0, 0, 0]
    for name, region, hc, jul, aug, sep, vs, status in depots:
        q3 = jul + aug + sep
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=region)
        ws.cell(row=r, column=3, value=hc)
        ws.cell(row=r, column=4, value=jul)
        ws.cell(row=r, column=5, value=aug)
        ws.cell(row=r, column=6, value=sep)
        ws.cell(row=r, column=7, value=q3)
        ws.cell(row=r, column=8, value=round(q3 / hc))
        ws.cell(row=r, column=9, value=vs)
        ws.cell(row=r, column=10, value=status)
        for i, v in enumerate((hc, jul, aug, sep, q3)):
            tot[i] += v
        r += 1
    ws.cell(row=r, column=1, value="All depots").font = Font(bold=True)
    for i, v in enumerate(tot):
        ws.cell(row=r, column=3 + i, value=v).font = Font(bold=True)
    ws.cell(row=r, column=8, value=round(tot[4] / tot[0])).font = Font(bold=True)

    for rr in range(4, r + 1):
        for c in range(1, 11):
            cell = ws.cell(row=rr, column=c)
            cell.border = BORDER
            if 4 <= c <= 8:
                cell.number_format = MONEY
            if c == 9:
                cell.number_format = "+0.0%;-0.0%"
            if c == 3:
                cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 11
    for c in range(3, 11):
        ws.column_dimensions[get_column_letter(c)].width = 13


def headcount_sheet(wb):
    ws = wb.create_sheet("Headcount")
    ws["A1"] = "Establishment and vacancies"
    ws["A1"].font = Font(bold=True, size=13, color="1F3A5F")
    hdr = ["Role", "Grade", "Budgeted FTE", "Filled", "Vacant", "Avg salary",
           "On-cost %", "Fully loaded", "Notes"]
    ws.append([])
    ws.append(hdr)
    header_row(ws, 3, len(hdr))
    roles = [
        ("Field technician", "T3", 62, 57, 31400, 0.212, "3 offers out"),
        ("Senior field technician", "T4", 24, 24, 38900, 0.212, ""),
        ("Depot supervisor", "M2", 11, 10, 44700, 0.223, "Willowmere open"),
        ("Plant operator", "T3", 18, 16, 33200, 0.212, "Agency covering"),
        ("Scheduler", "A2", 9, 9, 29800, 0.205, ""),
        ("Compliance officer", "P2", 5, 4, 41600, 0.218, "Interviewing"),
        ("Fleet coordinator", "A3", 4, 4, 32500, 0.205, ""),
        ("Stores keeper", "A2", 8, 7, 27900, 0.205, ""),
        ("HSE advisor", "P2", 3, 3, 45300, 0.218, ""),
        ("Contract manager", "M3", 6, 5, 56800, 0.231, "Backfill in Q4"),
        ("Data analyst", "P1", 3, 2, 38200, 0.212, "New post for Q3"),
        ("Apprentice", "T1", 12, 11, 19600, 0.198, "September intake"),
        ("Ecologist", "P3", 4, 4, 43100, 0.218, ""),
        ("Surveyor", "P2", 7, 6, 39800, 0.218, "1 on secondment"),
        ("Administrator", "A1", 6, 6, 25400, 0.205, ""),
    ]
    r = 4
    for role, grade, fte, filled, sal, onc, note in roles:
        ws.cell(row=r, column=1, value=role)
        ws.cell(row=r, column=2, value=grade)
        ws.cell(row=r, column=3, value=fte)
        ws.cell(row=r, column=4, value=filled)
        ws.cell(row=r, column=5, value=fte - filled)
        ws.cell(row=r, column=6, value=sal)
        ws.cell(row=r, column=7, value=onc)
        ws.cell(row=r, column=8, value=round(filled * sal * (1 + onc)))
        ws.cell(row=r, column=9, value=note)
        r += 1
    for rr in range(4, r):
        for c in range(1, 10):
            cell = ws.cell(row=rr, column=c)
            cell.border = BORDER
            if c in (6, 8):
                cell.number_format = MONEY
            if c == 7:
                cell.number_format = "0.0%"
            if c in (2, 3, 4, 5):
                cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["I"].width = 20
    for c in range(2, 9):
        ws.column_dimensions[get_column_letter(c)].width = 12


def capex_sheet(wb):
    ws = wb.create_sheet("Capex")
    ws["A1"] = "Capital requests awaiting approval"
    ws["A1"].font = Font(bold=True, size=13, color="1F3A5F")
    hdr = ["Ref", "Item", "Depot", "Requested", "Approved", "Payback (months)",
           "Decision", "Sponsor"]
    ws.append([])
    ws.append(hdr)
    header_row(ws, 3, len(hdr))
    capex = [
        ("CX-2611", "Replace 3.5t tipper × 2", "Willowmere Yard", 74800, 74800, 41, "Approved", "R. Adeyemi"),
        ("CX-2612", "Mobile welfare unit", "Saltmarsh Access", 21400, 0, 28, "Deferred", "L. Brennan"),
        ("CX-2613", "Depot roof re-covering", "Marrow Lane Works", 96500, 88000, 0, "Approved", "M. Okonkwo"),
        ("CX-2614", "Telematics retrofit, 46 vehicles", "All", 38640, 38640, 14, "Approved", "T. Vasquez"),
        ("CX-2615", "Tool store racking", "Thornhill Row", 12900, 12900, 0, "Approved", "L. Brennan"),
        ("CX-2616", "Compressor overhaul", "Fenwick Sidings", 17250, 0, 22, "On hold", "R. Adeyemi"),
        ("CX-2617", "Site cameras, 4 compounds", "Various", 29300, 24000, 19, "Part approved", "S. Nakamura"),
        ("CX-2618", "EV charge points × 6", "Kestrel Bay Depot", 54100, 0, 63, "Under review", "T. Vasquez"),
        ("CX-2619", "Survey drone and licence", "Willowmere Yard", 8750, 8750, 11, "Approved", "S. Nakamura"),
        ("CX-2620", "Portable water treatment rig", "Greywater Pumping", 43900, 43900, 26, "Approved", "M. Okonkwo"),
        ("CX-2621", "Yard resurfacing, phase 1", "Hollowfield Compound", 61200, 30000, 0, "Part approved", "L. Brennan"),
    ]
    r = 4
    for row in capex:
        for i, v in enumerate(row, start=1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    ws.cell(row=r, column=2, value="Total").font = Font(bold=True)
    ws.cell(row=r, column=4, value=sum(x[3] for x in capex)).font = Font(bold=True)
    ws.cell(row=r, column=5, value=sum(x[4] for x in capex)).font = Font(bold=True)
    for rr in range(4, r + 1):
        for c in range(1, 9):
            cell = ws.cell(row=rr, column=c)
            cell.border = BORDER
            if c in (4, 5):
                cell.number_format = MONEY
            if c == 6:
                cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 21
    for c in range(4, 9):
        ws.column_dimensions[get_column_letter(c)].width = 13


def assumptions_sheet(wb):
    ws = wb.create_sheet("Assumptions")
    ws["A1"] = "Basis of preparation"
    ws["A1"].font = Font(bold=True, size=13, color="1F3A5F")
    ws.append([])
    ws.append(["Driver", "Value", "Source", "Last reviewed"])
    header_row(ws, 3, 4, height=22)
    assum = [
        ("Pay award, effective 1 July", "3.20%", "Board minute 26/07", "12 Jun 2026"),
        ("Employer pension contribution", "6.00%", "Scheme rules v4", "01 Apr 2026"),
        ("Employer NI rate", "13.80%", "Payroll bureau", "06 Apr 2026"),
        ("Diesel, pence per litre", "148.5", "Fuel card 3-month average", "31 Jul 2026"),
        ("Fleet utilisation", "78%", "Telematics export", "31 Jul 2026"),
        ("Chargeable hours per FTE", "1,512", "Operating model v9", "20 May 2026"),
        ("Agency uplift on base rate", "21%", "Framework schedule", "01 Jan 2026"),
        ("Inflation applied to consumables", "2.40%", "Procurement index", "30 Jun 2026"),
        ("Bad debt provision", "0.90%", "Ledger 12-month actual", "30 Jun 2026"),
        ("Contingency, % of opex", "3.00%", "Finance policy FP-11", "01 Apr 2026"),
        ("Depot overhead recovery", "17.5%", "Cost model 2026", "01 Apr 2026"),
        ("Plant hire day rate uplift", "4.10%", "Supplier notice", "14 Jul 2026"),
    ]
    r = 4
    for row in assum:
        for i, v in enumerate(row, start=1):
            ws.cell(row=r, column=i, value=v).border = BORDER
        r += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 16


if __name__ == "__main__":
    wb = Workbook()
    summary_sheet(wb)
    depots_sheet(wb)
    headcount_sheet(wb)
    capex_sheet(wb)
    assumptions_sheet(wb)
    path = OUT / "Q3 Operating Budget FY26.xlsx"
    wb.save(path)
    print("wrote", path, path.stat().st_size // 1024, "KB")
